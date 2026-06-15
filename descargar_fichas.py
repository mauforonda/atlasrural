#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import logging
import math
import random
import sqlite3
import threading
import time
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from tqdm import tqdm

API_BASE = "https://idg.ine.gob.bo/api"
DEFAULT_UNIDADES = Path("datos/comunidades.parquet")
DEFAULT_CAMPOS = Path("recursos/campos.json")
DEFAULT_DB = Path("temporal/fichas.sqlite")
DEFAULT_POBLACION = Path("datos/poblacion.parquet")
DEFAULT_FICHAS = Path("datos/fichas.parquet")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:144.0) Gecko/20100101 Firefox/144.0",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.5",
    "Content-Type": "application/json",
    "Origin": "https://idg.ine.gob.bo",
    "Referer": "https://idg.ine.gob.bo/geoportal",
}

LOGGER = logging.getLogger("descargar_fichas")
_thread_local = threading.local()


class TqdmLoggingHandler(logging.Handler):
    def emit(self, record: logging.LogRecord) -> None:
        try:
            tqdm.write(self.format(record))
            self.flush()
        except Exception:
            self.handleError(record)


def configure_logging(level: str) -> None:
    handler = TqdmLoggingHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logging.basicConfig(
        level=getattr(logging, level),
        handlers=[handler],
        force=True,
    )


def qident(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


class TokenManager:
    def __init__(self, timeout: float):
        self.timeout = timeout
        self._token: str | None = None
        self._lock = threading.Lock()

    def get_token(self, force_refresh: bool = False) -> str:
        with self._lock:
            if self._token is None or force_refresh:
                self._token = self._fetch_token()
            return self._token

    def _fetch_token(self) -> str:
        response = requests.post(
            f"{API_BASE}/auth/acceso",
            headers=HEADERS,
            json={},
            timeout=self.timeout,
        )
        response.raise_for_status()
        token = response.json().get("token")
        if not token:
            raise RuntimeError("auth/acceso did not return token")
        return token


def get_session() -> requests.Session:
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=4, pool_maxsize=4)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        _thread_local.session = session
    return session


def get_db_connection(db_path: Path) -> sqlite3.Connection:
    connection = getattr(_thread_local, "db_connection", None)
    if connection is None:
        connection = sqlite3.connect(db_path, timeout=30, check_same_thread=False)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA synchronous=NORMAL")
        connection.execute("PRAGMA busy_timeout=30000")
        _thread_local.db_connection = connection
    return connection


def cell_value_to_json(value: Any) -> Any:
    if hasattr(value, "item"):
        return value.item()
    return value


def parse_field_value(worksheet, cell_ref: str) -> Any:
    value = worksheet[cell_ref].value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return cell_value_to_json(value)

    row = worksheet[cell_ref].row
    col = worksheet[cell_ref].column
    for next_col in range(col + 1, worksheet.max_column + 1):
        candidate = worksheet.cell(row=row, column=next_col).value
        if isinstance(candidate, (int, float)) and not isinstance(candidate, bool):
            return cell_value_to_json(candidate)

    return cell_value_to_json(value)


def parse_known_fields(workbook_bytes: bytes, campos: dict[str, str]) -> dict[str, Any]:
    worksheet = load_workbook(
        BytesIO(workbook_bytes), data_only=True, read_only=True
    ).active
    return {
        field_name: parse_field_value(worksheet, cell_ref)
        for cell_ref, field_name in campos.items()
    }


def build_headers(token: str) -> dict[str, str]:
    return {**HEADERS, "Authorization": f"Bearer {token}"}


def content_type(response: requests.Response) -> str:
    return (response.headers.get("content-type") or "").lower()


def request_json(
    session: requests.Session,
    token_manager: TokenManager,
    path: str,
    payload: dict[str, Any],
    timeout: float,
    max_retries: int,
) -> dict[str, Any]:
    backoff = 1.0
    last_error: Exception | None = None

    for attempt in range(1, max_retries + 1):
        token = token_manager.get_token(force_refresh=False)
        try:
            response = session.post(
                f"{API_BASE}{path}",
                headers=build_headers(token),
                json=payload,
                timeout=timeout,
            )
        except requests.RequestException as exc:
            last_error = exc
        else:
            if response.status_code == 200:
                try:
                    return response.json()
                except Exception as exc:
                    last_error = exc
            elif response.status_code in (401, 403):
                token_manager.get_token(force_refresh=True)
            elif response.status_code in (429, 500, 502, 503, 504):
                last_error = RuntimeError(
                    f"{path} returned {response.status_code}: {response.text[:200]}"
                )
            else:
                response.raise_for_status()

        if attempt < max_retries:
            time.sleep(backoff + random.random() * 0.5)
            backoff = min(backoff * 2, 60)

    if last_error:
        raise last_error
    raise RuntimeError(f"Request failed without explicit error: {path}")


def request_excel(
    session: requests.Session,
    token_manager: TokenManager,
    payload: dict[str, Any],
    timeout: float,
    max_retries: int,
) -> bytes:
    backoff = 1.0
    last_error: Exception | None = None
    accepted_types = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    )

    for attempt in range(1, max_retries + 1):
        token = token_manager.get_token(force_refresh=False)
        try:
            response = session.post(
                f"{API_BASE}/ficha/generar-excel",
                headers=build_headers(token),
                json=payload,
                timeout=timeout,
            )
        except requests.RequestException as exc:
            last_error = exc
        else:
            ct = content_type(response)
            if response.status_code == 200 and any(t in ct for t in accepted_types):
                return response.content
            if response.status_code in (401, 403):
                token_manager.get_token(force_refresh=True)
            elif response.status_code in (429, 500, 502, 503, 504):
                last_error = RuntimeError(
                    f"generar-excel returned {response.status_code}: {response.text[:200]}"
                )
            else:
                if response.status_code == 200:
                    preview = response.text[:200]
                    last_error = RuntimeError(
                        f"Unexpected content-type {ct!r} from generar-excel: {preview}"
                    )
                else:
                    response.raise_for_status()

        if attempt < max_retries:
            time.sleep(backoff + random.random() * 0.5)
            backoff = min(backoff * 2, 60)

    if last_error:
        raise last_error
    raise RuntimeError("Excel request failed without explicit error")


def load_campos(campos_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    with open(campos_path, "r", encoding="utf-8") as f:
        campos = json.load(f)
    if not isinstance(campos, dict) or "base" not in campos or "vivienda" not in campos:
        raise RuntimeError("campos.json debe tener llaves 'base' y 'vivienda'")
    return campos["base"], campos["vivienda"]


def dedupe_keep_order(values: list[str]) -> list[str]:
    return list(dict.fromkeys(values))


def json_scalar(value: Any) -> Any:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, float) and math.isnan(value):
        return None
    return value


def init_database(db_path: Path, ficha_columns: list[str]) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(db_path, timeout=30)
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA busy_timeout=30000")

    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS progress (
            codigo TEXT PRIMARY KEY,
            verify_done INTEGER NOT NULL DEFAULT 0,
            validado INTEGER,
            personas INTEGER,
            viviendas INTEGER,
            mensaje TEXT,
            base_done INTEGER NOT NULL DEFAULT 0,
            vivienda_done INTEGER NOT NULL DEFAULT 0,
            completed INTEGER NOT NULL DEFAULT 0,
            error_count INTEGER NOT NULL DEFAULT 0,
            last_error TEXT,
            updated_at TEXT
        )
        """
    )
    connection.execute(
        """
        CREATE TABLE IF NOT EXISTS poblacion (
            codigo TEXT PRIMARY KEY,
            validado INTEGER,
            personas INTEGER,
            viviendas INTEGER
        )
        """
    )
    connection.execute(
        f"""
        CREATE TABLE IF NOT EXISTS fichas (
            codigo TEXT PRIMARY KEY,
            {", ".join(f"{qident(column)} REAL" for column in ficha_columns)}
        )
        """
    )

    existing_columns = {
        row[1] for row in connection.execute("PRAGMA table_info(fichas)").fetchall()
    }
    for column in ficha_columns:
        if column not in existing_columns:
            connection.execute(
                f"ALTER TABLE fichas ADD COLUMN {qident(column)} REAL"
            )

    connection.commit()
    connection.close()


def upsert_poblacion_rows(
    connection: sqlite3.Connection,
    rows: list[tuple[str, int | None, int | None, int | None]],
) -> None:
    if not rows:
        return

    now = utc_now()
    connection.executemany(
        """
        INSERT INTO poblacion (codigo, validado, personas, viviendas)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            validado = excluded.validado,
            personas = excluded.personas,
            viviendas = excluded.viviendas
        """,
        rows,
    )
    connection.executemany(
        """
        INSERT INTO progress (
            codigo, verify_done, validado, personas, viviendas, updated_at
        ) VALUES (?, 1, ?, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            verify_done = 1,
            validado = COALESCE(progress.validado, excluded.validado),
            personas = COALESCE(progress.personas, excluded.personas),
            viviendas = COALESCE(progress.viviendas, excluded.viviendas),
            updated_at = excluded.updated_at
        """,
        [(codigo, validado, personas, viviendas, now) for codigo, validado, personas, viviendas in rows],
    )
    connection.commit()


def upsert_ficha_chunk(
    connection: sqlite3.Connection,
    rows: list[dict[str, Any]],
    stage_columns: dict[str, int],
) -> None:
    if not rows:
        return

    columns = ["codigo"] + [column for column in rows[0].keys() if column != "codigo"]
    insert_columns = ", ".join(qident(column) for column in columns)
    placeholders = ", ".join(["?"] * len(columns))
    update_clause = ", ".join(
        f"{qident(column)} = COALESCE(excluded.{qident(column)}, fichas.{qident(column)})"
        for column in columns
        if column != "codigo"
    )
    values = [[row.get(column) for column in columns] for row in rows]
    now = utc_now()

    connection.executemany(
        f"""
        INSERT INTO fichas ({insert_columns})
        VALUES ({placeholders})
        ON CONFLICT(codigo) DO UPDATE SET
            {update_clause}
        """,
        values,
    )
    connection.executemany(
        """
        INSERT INTO progress (
            codigo, verify_done, validado, base_done, vivienda_done, updated_at
        ) VALUES (?, 1, 1, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            verify_done = 1,
            validado = 1,
            base_done = MAX(progress.base_done, excluded.base_done),
            vivienda_done = MAX(progress.vivienda_done, excluded.vivienda_done),
            updated_at = excluded.updated_at
        """,
        [
            (
                row["codigo"],
                stage_columns["base_done"],
                stage_columns["vivienda_done"],
                now,
            )
            for row in rows
        ],
    )
    connection.commit()


def seed_from_existing(
    db_path: Path,
    poblacion_path: Path | None,
    fichas_path: Path | None,
    campos_base: dict[str, str],
    campos_vivienda: dict[str, str],
    chunksize: int,
    trust_old_invalid: bool,
) -> None:
    connection = sqlite3.connect(db_path, timeout=30)
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    connection.execute("PRAGMA busy_timeout=30000")

    if poblacion_path is not None and poblacion_path.exists():
        LOGGER.info("Sembrando población desde %s", poblacion_path)
        parquet = pq.ParquetFile(poblacion_path)
        for batch in parquet.iter_batches(
            columns=["codigo", "validado", "personas", "viviendas"],
            batch_size=chunksize,
        ):
            frame = batch.to_pandas()
            rows = []
            for row in frame.itertuples(index=False):
                validado = json_scalar(row.validado)
                personas = json_scalar(row.personas)
                viviendas = json_scalar(row.viviendas)
                rows.append(
                    (
                        str(row.codigo),
                        None if validado is None else int(bool(validado)),
                        None if personas is None else int(personas),
                        None if viviendas is None else int(viviendas),
                    )
                )
            upsert_poblacion_rows(connection, rows)

    if fichas_path is not None and fichas_path.exists():
        LOGGER.info("Sembrando fichas desde %s", fichas_path)
        parquet = pq.ParquetFile(fichas_path)
        available_columns = set(parquet.schema.names)
        selected_base = [column for column in campos_base.values() if column in available_columns]
        selected_vivienda = [
            column for column in campos_vivienda.values() if column in available_columns
        ]
        selected_columns = ["codigo"] + dedupe_keep_order(selected_base + selected_vivienda)
        stage_columns = {
            "base_done": int(bool(selected_base)),
            "vivienda_done": int(
                bool(selected_vivienda)
                and len(selected_vivienda) == len(set(campos_vivienda.values()))
            ),
        }
        for batch in parquet.iter_batches(columns=selected_columns, batch_size=chunksize):
            frame = batch.to_pandas()
            rows = []
            for item in frame.to_dict(orient="records"):
                rows.append(
                    {
                        column: json_scalar(value)
                        for column, value in item.items()
                        if column == "codigo" or value is not None
                    }
                )
            upsert_ficha_chunk(connection, rows, stage_columns)

    if trust_old_invalid:
        LOGGER.info("Marcando como completados los casos inválidos ya presentes en población")
        connection.execute(
            """
            UPDATE progress
            SET completed = 1,
                updated_at = ?,
                last_error = NULL
            WHERE verify_done = 1
              AND COALESCE(validado, 0) = 0
            """,
            (utc_now(),),
        )
        connection.commit()

    connection.close()


def fetch_progress_row(connection: sqlite3.Connection, codigo: str) -> dict[str, Any]:
    row = connection.execute(
        "SELECT * FROM progress WHERE codigo = ?",
        (codigo,),
    ).fetchone()
    if row is None:
        return {
            "codigo": codigo,
            "verify_done": 0,
            "validado": None,
            "personas": None,
            "viviendas": None,
            "mensaje": None,
            "base_done": 0,
            "vivienda_done": 0,
            "completed": 0,
            "error_count": 0,
            "last_error": None,
            "updated_at": None,
        }
    return dict(row)


def save_summary(
    connection: sqlite3.Connection,
    codigo: str,
    verification: dict[str, Any],
) -> None:
    validado = int(bool(verification.get("validado", False)))
    personas = verification.get("cantidad_personas")
    viviendas = verification.get("cantidad_viviendas")
    mensaje = verification.get("mensaje")
    now = utc_now()

    connection.execute(
        """
        INSERT INTO progress (
            codigo, verify_done, validado, personas, viviendas, mensaje,
            updated_at, last_error
        ) VALUES (?, 1, ?, ?, ?, ?, ?, NULL)
        ON CONFLICT(codigo) DO UPDATE SET
            verify_done = 1,
            validado = excluded.validado,
            personas = excluded.personas,
            viviendas = excluded.viviendas,
            mensaje = excluded.mensaje,
            updated_at = excluded.updated_at,
            last_error = NULL
        """,
        (codigo, validado, personas, viviendas, mensaje, now),
    )
    connection.execute(
        """
        INSERT INTO poblacion (codigo, validado, personas, viviendas)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            validado = excluded.validado,
            personas = excluded.personas,
            viviendas = excluded.viviendas
        """,
        (codigo, validado, personas, viviendas),
    )
    connection.commit()


def save_ficha_fields(
    connection: sqlite3.Connection,
    codigo: str,
    fields: dict[str, Any],
    stage: str,
) -> None:
    if not fields:
        return

    columns = ["codigo"] + list(fields.keys())
    values = [codigo] + [fields[column] for column in fields.keys()]

    insert_columns = ", ".join(qident(column) for column in columns)
    placeholders = ", ".join(["?"] * len(columns))
    update_clause = ", ".join(
        f"{qident(column)} = excluded.{qident(column)}" for column in fields.keys()
    )
    now = utc_now()
    stage_column = "base_done" if stage == "base" else "vivienda_done"

    connection.execute(
        f"""
        INSERT INTO fichas ({insert_columns})
        VALUES ({placeholders})
        ON CONFLICT(codigo) DO UPDATE SET
            {update_clause}
        """,
        values,
    )
    connection.execute(
        f"""
        INSERT INTO progress (codigo, {stage_column}, updated_at, last_error)
        VALUES (?, 1, ?, NULL)
        ON CONFLICT(codigo) DO UPDATE SET
            {stage_column} = 1,
            updated_at = excluded.updated_at,
            last_error = NULL
        """,
        (codigo, now),
    )
    connection.commit()


def mark_completed(connection: sqlite3.Connection, codigo: str) -> None:
    connection.execute(
        """
        INSERT INTO progress (codigo, completed, updated_at, last_error)
        VALUES (?, 1, ?, NULL)
        ON CONFLICT(codigo) DO UPDATE SET
            completed = 1,
            updated_at = excluded.updated_at,
            last_error = NULL
        """,
        (codigo, utc_now()),
    )
    connection.commit()


def record_error(
    connection: sqlite3.Connection,
    codigo: str,
    stage: str,
    exc: Exception,
) -> None:
    message = f"{stage}: {type(exc).__name__}: {exc}"
    connection.execute(
        """
        INSERT INTO progress (codigo, error_count, last_error, updated_at)
        VALUES (?, 1, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            error_count = progress.error_count + 1,
            last_error = excluded.last_error,
            updated_at = excluded.updated_at
        """,
        (codigo, message[:2000], utc_now()),
    )
    connection.commit()


def process_code(
    codigo: str,
    db_path: Path,
    token_manager: TokenManager,
    campos_base: dict[str, str],
    campos_vivienda: dict[str, str],
    verify_timeout: float,
    excel_timeout: float,
    max_retries: int,
) -> dict[str, Any]:
    connection = get_db_connection(db_path)
    session = get_session()
    state = fetch_progress_row(connection, codigo)

    try:
        if not state["verify_done"]:
            verification = request_json(
                session=session,
                token_manager=token_manager,
                path="/ficha/verificarValidar",
                payload={"mara": 2024, "codigos": [codigo]},
                timeout=verify_timeout,
                max_retries=max_retries,
            )
            save_summary(connection, codigo, verification)
            state = fetch_progress_row(connection, codigo)

        validado = bool(state["validado"])
        if not validado:
            if not state["completed"]:
                mark_completed(connection, codigo)
            return {"codigo": codigo, "validado": False}

        if not state["base_done"]:
            base_bytes = request_excel(
                session=session,
                token_manager=token_manager,
                payload={"mara": "2024", "codigos": [codigo], "vivienda": False},
                timeout=excel_timeout,
                max_retries=max_retries,
            )
            base_fields = parse_known_fields(base_bytes, campos_base)
            save_ficha_fields(connection, codigo, base_fields, stage="base")
            state = fetch_progress_row(connection, codigo)

        if not state["vivienda_done"]:
            vivienda_bytes = request_excel(
                session=session,
                token_manager=token_manager,
                payload={"mara": "2024", "codigos": [codigo], "vivienda": True},
                timeout=excel_timeout,
                max_retries=max_retries,
            )
            vivienda_fields = parse_known_fields(vivienda_bytes, campos_vivienda)
            save_ficha_fields(connection, codigo, vivienda_fields, stage="vivienda")

        if not state["completed"]:
            mark_completed(connection, codigo)

        return {"codigo": codigo, "validado": True}
    except Exception as exc:
        stage = "unknown"
        if not state["verify_done"]:
            stage = "verify"
        elif not state["base_done"]:
            stage = "base"
        elif not state["vivienda_done"]:
            stage = "vivienda"
        record_error(connection, codigo, stage, exc)
        raise


def count_progress(connection: sqlite3.Connection, codes: list[str]) -> dict[str, int]:
    code_set = set(codes)
    completed = 0
    validos = 0
    invalidos = 0
    con_error = 0

    for row in connection.execute(
        "SELECT codigo, completed, validado, last_error FROM progress"
    ).fetchall():
        codigo = row["codigo"]
        if codigo not in code_set:
            continue
        if row["completed"]:
            completed += 1
            if row["validado"] == 1:
                validos += 1
            elif row["validado"] == 0:
                invalidos += 1
        elif row["last_error"] is not None:
            con_error += 1

    return {
        "completed": completed,
        "validos": validos,
        "invalidos": invalidos,
        "con_error": con_error,
    }


def load_codes(unidades_path: Path, limit: int | None = None) -> list[str]:
    series = pd.read_parquet(unidades_path, columns=["codigo"])["codigo"].astype(str)
    codes = series.drop_duplicates().tolist()
    if limit is not None:
        codes = codes[:limit]
    return codes


def pending_codes(connection: sqlite3.Connection, codes: list[str]) -> list[str]:
    completed = {
        row["codigo"]
        for row in connection.execute(
            "SELECT codigo FROM progress WHERE completed = 1"
        ).fetchall()
    }
    return [code for code in codes if code not in completed]


def atomic_parquet_write(
    query: str,
    output_path: Path,
    connection: sqlite3.Connection,
    empty_columns: list[str],
    chunksize: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_name(output_path.name + ".tmp")
    if tmp_path.exists():
        tmp_path.unlink()

    writer: pq.ParquetWriter | None = None
    wrote_rows = False

    try:
        for chunk in pd.read_sql_query(query, connection, chunksize=chunksize):
            if chunk.empty:
                continue
            wrote_rows = True
            table = pa.Table.from_pandas(chunk, preserve_index=False)
            if writer is None:
                writer = pq.ParquetWriter(tmp_path, table.schema)
            writer.write_table(table)

        if writer is not None:
            writer.close()
            writer = None
        elif not wrote_rows:
            pd.DataFrame(columns=empty_columns).to_parquet(tmp_path, index=False)

        tmp_path.replace(output_path)
    finally:
        if writer is not None:
            writer.close()
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def export_parquets(
    db_path: Path,
    poblacion_path: Path,
    fichas_path: Path,
    codes: list[str],
    ficha_columns: list[str],
    chunksize: int,
) -> None:
    connection = sqlite3.connect(db_path, timeout=30)
    connection.execute("DROP TABLE IF EXISTS selected_codes")
    connection.execute("CREATE TEMP TABLE selected_codes (codigo TEXT PRIMARY KEY)")
    connection.executemany(
        "INSERT OR IGNORE INTO selected_codes (codigo) VALUES (?)",
        [(code,) for code in codes],
    )

    poblacion_query = """
        SELECT codigo, CAST(validado AS BOOLEAN) AS validado, personas, viviendas
        FROM poblacion
        WHERE codigo IN (SELECT codigo FROM selected_codes)
        ORDER BY codigo
    """
    fichas_query = f"""
        SELECT {", ".join([qident('codigo')] + [qident(column) for column in ficha_columns])}
        FROM fichas
        WHERE codigo IN (
            SELECT codigo
            FROM progress
            WHERE completed = 1 AND validado = 1
        )
          AND codigo IN (SELECT codigo FROM selected_codes)
        ORDER BY codigo
    """

    LOGGER.info("Exportando %s", poblacion_path)
    atomic_parquet_write(
        poblacion_query,
        poblacion_path,
        connection,
        ["codigo", "validado", "personas", "viviendas"],
        chunksize,
    )
    LOGGER.info("Exportando %s", fichas_path)
    atomic_parquet_write(
        fichas_query,
        fichas_path,
        connection,
        ["codigo"] + ficha_columns,
        chunksize,
    )
    connection.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Descarga fichas censales de manera resiliente y exporta parquet."
    )
    parser.add_argument(
        "--mode",
        choices=["download", "export", "download-export"],
        default="download-export",
        help="Descargar, exportar o hacer ambas cosas.",
    )
    parser.add_argument(
        "--unidades",
        type=Path,
        default=DEFAULT_UNIDADES,
        help="Parquet con códigos de comunidad.",
    )
    parser.add_argument(
        "--campos",
        type=Path,
        default=DEFAULT_CAMPOS,
        help="JSON con mappings de campos base y vivienda.",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=DEFAULT_DB,
        help="SQLite local de trabajo y reanudación.",
    )
    parser.add_argument(
        "--poblacion-output",
        type=Path,
        default=DEFAULT_POBLACION,
        help="Parquet final para población y vivienda.",
    )
    parser.add_argument(
        "--fichas-output",
        type=Path,
        default=DEFAULT_FICHAS,
        help="Parquet final para fichas base + vivienda.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=2,
        help="Cantidad de comunidades a procesar en paralelo.",
    )
    parser.add_argument(
        "--verify-timeout",
        type=float,
        default=30.0,
        help="Timeout para verificar validación.",
    )
    parser.add_argument(
        "--excel-timeout",
        type=float,
        default=120.0,
        help="Timeout para descargar cada Excel.",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=7,
        help="Intentos máximos por request.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Límite de comunidades para pruebas.",
    )
    parser.add_argument(
        "--export-chunksize",
        type=int,
        default=5000,
        help="Tamaño de chunk para exportar parquet.",
    )
    parser.add_argument(
        "--seed-poblacion",
        type=Path,
        default=DEFAULT_POBLACION,
        help="Parquet existente para sembrar población y evitar requests repetidos.",
    )
    parser.add_argument(
        "--seed-fichas",
        type=Path,
        default=DEFAULT_FICHAS,
        help="Parquet existente para sembrar fichas previas y evitar redescargar base.",
    )
    parser.add_argument(
        "--seed-chunksize",
        type=int,
        default=5000,
        help="Tamaño de chunk para sembrar desde parquet existente.",
    )
    parser.add_argument(
        "--trust-old-invalid",
        action="store_true",
        help="Si ya existe población con validado=false, marcarlo como completado sin consultar la API.",
    )
    parser.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default="INFO",
    )
    return parser.parse_args()


def run_download(args: argparse.Namespace) -> None:
    campos_base, campos_vivienda = load_campos(args.campos)
    ficha_columns = dedupe_keep_order(
        list(campos_base.values()) + list(campos_vivienda.values())
    )
    init_database(args.db, ficha_columns)
    seed_from_existing(
        db_path=args.db,
        poblacion_path=args.seed_poblacion,
        fichas_path=args.seed_fichas,
        campos_base=campos_base,
        campos_vivienda=campos_vivienda,
        chunksize=args.seed_chunksize,
        trust_old_invalid=args.trust_old_invalid,
    )

    codes = load_codes(args.unidades, limit=args.limit)
    db_connection = sqlite3.connect(args.db, timeout=30)
    db_connection.row_factory = sqlite3.Row
    initial = count_progress(db_connection, codes)
    pending = pending_codes(db_connection, codes)
    db_connection.close()

    if not pending:
        LOGGER.info(
            "No hay comunidades pendientes. Completados=%s validos=%s invalidos=%s",
            initial["completed"],
            initial["validos"],
            initial["invalidos"],
        )
        return

    token_manager = TokenManager(timeout=args.verify_timeout)
    total = len(codes)
    completed = initial["completed"]
    validos = initial["validos"]
    invalidos = initial["invalidos"]
    errors = initial["con_error"]

    LOGGER.info(
        "Procesando %s comunidades pendientes de %s totales con %s workers",
        len(pending),
        total,
        args.workers,
    )

    future_to_code: dict[Any, str] = {}
    pending_iter = iter(pending)

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        def submit_next() -> bool:
            try:
                code = next(pending_iter)
            except StopIteration:
                return False
            future = executor.submit(
                process_code,
                code,
                args.db,
                token_manager,
                campos_base,
                campos_vivienda,
                args.verify_timeout,
                args.excel_timeout,
                args.max_retries,
            )
            future_to_code[future] = code
            return True

        for _ in range(min(args.workers, len(pending))):
            submit_next()

        with tqdm(
            total=total,
            initial=completed,
            desc="Comunidades",
            unit="comunidad",
        ) as progress:
            progress.set_postfix(validos=validos, invalidos=invalidos, errores=errors)
            try:
                while future_to_code:
                    done, _ = wait(future_to_code, return_when=FIRST_COMPLETED)
                    for future in done:
                        code = future_to_code.pop(future)
                        try:
                            result = future.result()
                        except Exception as exc:
                            errors += 1
                            LOGGER.warning("Fallo %s: %s", code, exc)
                        else:
                            completed += 1
                            if result["validado"]:
                                validos += 1
                            else:
                                invalidos += 1
                            progress.update(1)
                            progress.set_postfix(
                                validos=validos,
                                invalidos=invalidos,
                                errores=errors,
                            )
                        finally:
                            submit_next()
            except KeyboardInterrupt:
                LOGGER.warning(
                    "Interrupción recibida. El progreso ya persistido queda en %s",
                    args.db,
                )
                executor.shutdown(wait=False, cancel_futures=True)
                raise


def main() -> None:
    args = parse_args()
    configure_logging(args.log_level)

    campos_base, campos_vivienda = load_campos(args.campos)
    ficha_columns = dedupe_keep_order(
        list(campos_base.values()) + list(campos_vivienda.values())
    )
    codes = load_codes(args.unidades, limit=args.limit)

    if args.mode in ("download", "download-export"):
        run_download(args)

    if args.mode in ("export", "download-export"):
        init_database(args.db, ficha_columns)
        export_parquets(
            db_path=args.db,
            poblacion_path=args.poblacion_output,
            fichas_path=args.fichas_output,
            codes=codes,
            ficha_columns=ficha_columns,
            chunksize=args.export_chunksize,
        )
        LOGGER.info("Exportación terminada")


if __name__ == "__main__":
    main()
